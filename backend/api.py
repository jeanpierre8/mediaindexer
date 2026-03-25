"""
api.py — Backend FastAPI pour Media Indexer
"""
import re, json, time, ctypes, string, logging, sqlite3, threading, requests, uvicorn, openpyxl, hashlib, shutil, uuid
from pathlib import Path
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

DOWNLOADS    = Path.home() / "Downloads"
DB_FILE      = Path.home() / "Documents" / "media_indexer.db"
LOG_FILE     = Path.home() / "Documents" / "media_indexer_api.log"
WATCHED_FILE = Path.home() / "Documents" / "media_indexer_watched.json"
CONFIG_FILE  = Path.home() / "Documents" / "media_indexer_config.json"
BACKUP_DIR   = Path.home() / "Documents" / "media_indexer_backups"
RULES_FILE   = Path.home() / "Documents" / "media_indexer_rules.json"
OLLAMA_URL   = "http://localhost:11434/api/generate"
TMDB_URL     = "https://api.themoviedb.org/3"
MODELE       = "qwen3.5:9b"

EXT_VIDEO  = {".mp4",".mkv",".avi",".mov",".wmv",".m4v",".ts",".flv",".webm"}
EXT_PDF    = {".pdf"}
EXT_IMAGE  = {".jpg",".jpeg",".png",".gif",".webp",".bmp",".tiff",".heic"}
EXT_AUDIO  = {".mp3",".flac",".aac",".wav",".ogg",".m4a",".wma",".opus"}
EXT_AUTRES = {".txt",".md",".csv",".docx",".xlsx",".zip",".rar",".7z"}
TOUTES_EXT = EXT_VIDEO|EXT_PDF|EXT_IMAGE|EXT_AUDIO|EXT_AUTRES

IGNORES = {"windows","program files","program files (x86)","programdata",
           "appdata","system volume information","$recycle.bin","recovery",
           ".git","__pycache__","node_modules"}

RE_QUALITE = [
    (re.compile(r'\b4[Kk]\b|\b2160[pP]\b'), '4K'),
    (re.compile(r'\b1080[pPiI]\b'), '1080p'),
    (re.compile(r'\b720[pPiI]\b'), '720p'),
    (re.compile(r'\b480[pPiI]\b'), '480p'),
    (re.compile(r'\bHDR\b|\bHDR10\b|\bDolbyVision\b|\bDV\b', re.I), 'HDR'),
    (re.compile(r'\bBluRay\b|\bBDRip\b|\bBD\b', re.I), 'BluRay'),
    (re.compile(r'\bWEBDL\b|\bWEB-DL\b|\bWEBRip\b', re.I), 'WEB'),
]

def detecter_qualite(nom):
    for pattern, label in RE_QUALITE:
        if pattern.search(nom): return label
    return ''

LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()])
log = logging.getLogger(__name__)

# ── État global ───────────────────────────────────────────────────
pending_disk    = None
watched_folders = []
observers       = []
scan_progress   = {"running": False, "current": 0, "total": 0, "folder": "", "done": False, "count": 0}
autotag_progress = {"running": False, "done": False, "folder": "", "current": "", "total": 0, "processed": 0, "tagged": 0, "skipped": 0, "errors": 0, "last_error": "", "overwrite": False, "only_untagged": True}
autotag_stop_flag = False

# ── Config ────────────────────────────────────────────────────────
def load_config():
    if CONFIG_FILE.exists():
        try: return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except: pass
    return {}

def save_config(data):
    CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def get_tmdb_key():        return load_config().get("tmdb_key", "")
def get_pin_hash():        return load_config().get("pin_hash", "")
def get_private_folders(): return load_config().get("private_folders", [])

def is_private(chemin: str) -> bool:
    return any(str(chemin).startswith(f) for f in get_private_folders())

def verify_pin(pin: str) -> bool:
    stored = get_pin_hash()
    if not stored: return False
    return hashlib.sha256(pin.encode()).hexdigest() == stored

# ── Règles ────────────────────────────────────────────────────────
def load_rules():
    if RULES_FILE.exists():
        try: return json.loads(RULES_FILE.read_text(encoding="utf-8"))
        except: pass
    return []

def save_rules(rules):
    RULES_FILE.write_text(json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Backup ────────────────────────────────────────────────────────
def do_backup():
    c = load_config()
    if not c.get("backup_enabled", True): return None
    if not DB_FILE.exists(): return None
    backup_dir = Path(c.get("backup_dir", str(BACKUP_DIR)))
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = backup_dir / f"media_indexer_{ts}.db"
    shutil.copy2(DB_FILE, dest)
    log.info(f"Backup : {dest.name}")
    keep = c.get("backup_keep", 7)
    backups = sorted(backup_dir.glob("media_indexer_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in backups[keep:]:
        old.unlink()
    return str(dest)

def backup_loop():
    while True:
        time.sleep(86400)
        try: do_backup()
        except Exception as e: log.error(f"Backup: {e}")

# ── Helpers ───────────────────────────────────────────────────────
RE_EP = [
    re.compile(r"[Ss](\d{1,2})[Ee](\d{1,3})"),
    re.compile(r"(\d{1,2})x(\d{1,3})"),
    re.compile(r"[Ss]eason\s*(\d+).*[Ee]p?\s*(\d+)", re.I),
    re.compile(r"[Ee](\d{2,3})\b"),
]

def detecter_serie(nom):
    stem  = Path(nom).stem
    clean = re.sub(r"[._]", " ", stem)
    for p in RE_EP:
        m = p.search(clean)
        if m:
            g = m.groups()
            s, e = (g[0], g[1]) if len(g)>=2 else ("1", g[0])
            serie = re.sub(r"[\[\(].*","", clean[:m.start()]).strip().title() or stem[:20]
            return serie, int(s), int(e)
    return None, None, None

def type_f(chemin):
    ext = Path(chemin).suffix.lower()
    if ext in EXT_VIDEO:  return "Vidéo"
    if ext in EXT_PDF:    return "PDF"
    if ext in EXT_IMAGE:  return "Image"
    if ext in EXT_AUDIO:  return "Musique"
    return "Autre"

# ── DB ────────────────────────────────────────────────────────────
def get_db():
    con = sqlite3.connect(DB_FILE)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    con = get_db()
    con.executescript("""
        CREATE TABLE IF NOT EXISTS fichiers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_scan TEXT, nom TEXT, type TEXT, disque TEXT,
            chemin TEXT UNIQUE, taille_mb REAL,
            serie TEXT, saison INTEGER, episode INTEGER,
            titre TEXT, resume TEXT, tags TEXT, analyse INTEGER DEFAULT 0,
            qualite TEXT DEFAULT '', vu INTEGER DEFAULT 0,
            favori INTEGER DEFAULT 0,
            tags_manuels TEXT DEFAULT '',
            note TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS series_meta (
            serie TEXT NOT NULL, saison INTEGER NOT NULL,
            total_ep INTEGER DEFAULT 0, tmdb_id INTEGER,
            source TEXT DEFAULT 'manuel', updated_at TEXT,
            PRIMARY KEY (serie, saison)
        );
        CREATE TABLE IF NOT EXISTS scan_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, dossier TEXT, nb_ajoutes INTEGER, duree_sec REAL
        );
    """)
    # Migrations — ajouter colonnes si absentes
    for col, typ, default in [
        ("qualite",      "TEXT",    "''"),
        ("vu",           "INTEGER", "0"),
        ("favori",       "INTEGER", "0"),
        ("tags_manuels", "TEXT",    "''"),
        ("note",         "TEXT",    "''"),
    ]:
        try:
            con.execute(f"ALTER TABLE fichiers ADD COLUMN {col} {typ} DEFAULT {default}")
            log.info(f"Migration : colonne {col} ajoutée")
        except: pass
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_type   ON fichiers(type)",
        "CREATE INDEX IF NOT EXISTS idx_serie  ON fichiers(serie)",
        "CREATE INDEX IF NOT EXISTS idx_chemin ON fichiers(chemin)",
        "CREATE INDEX IF NOT EXISTS idx_favori ON fichiers(favori)",
    ]:
        try: con.execute(idx)
        except: pass
    con.commit(); con.close()

def insert_file(nom, tf, disque, chemin, taille, serie=None, saison=None, episode=None, titre=""):
    qualite = detecter_qualite(nom) if tf == "Vidéo" else ""
    try:
        con = get_db()
        con.execute("""INSERT OR IGNORE INTO fichiers
            (date_scan,nom,type,disque,chemin,taille_mb,serie,saison,episode,titre,qualite)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (datetime.now().strftime("%Y-%m-%d %H:%M"), nom, tf, disque,
             str(chemin), taille, serie, saison, episode, titre, qualite))
        con.commit(); con.close()
        if serie and saison:
            threading.Thread(target=fetch_tmdb_if_missing, args=(serie, saison), daemon=True).start()
    except Exception as e: log.error(f"DB: {e}")

# ── Règles automatiques ───────────────────────────────────────────
def apply_rule(rule: dict, file_path: Path) -> bool:
    if not rule.get("actif", True): return False
    ct, cv = rule["condition_type"], rule["condition_val"].lower()
    name = file_path.name.lower()
    ext  = file_path.suffix.lower()
    tf   = type_f(file_path)
    match = False
    if ct == "extension"     and ext == cv:          match = True
    elif ct == "type"        and tf.lower() == cv:   match = True
    elif ct == "nom_contient" and cv in name:         match = True
    if not match: return False
    dest_dir = Path(rule["destination"])
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / file_path.name
    try:
        if rule["action"] == "deplacer": shutil.move(str(file_path), str(dest))
        else:                            shutil.copy2(str(file_path), str(dest))
        taille = round(dest.stat().st_size/(1024*1024), 2)
        tfr = type_f(dest)
        s, sa, ep = (None,None,None)
        if tfr == "Vidéo": s, sa, ep = detecter_serie(dest.name)
        insert_file(dest.name, tfr, str(dest.anchor).rstrip("\\/"), dest, taille, s, sa, ep, dest.stem)
        log.info(f"Règle '{rule['nom']}' : {file_path.name} → {dest_dir}")
        return True
    except Exception as e:
        log.error(f"Règle '{rule['nom']}': {e}")
        return False

# ── Scan ──────────────────────────────────────────────────────────
def scan_folder_background(path_str: str):
    global scan_progress
    path   = Path(path_str)
    disque = str(path.anchor).rstrip("\\/")
    try:
        all_files = []
        for f in path.rglob("*"):
            try:
                if any(p.lower() in IGNORES for p in f.parts): continue
                if not f.is_file(): continue
                if f.suffix.lower() not in TOUTES_EXT: continue
                all_files.append(f)
            except: continue
        scan_progress["total"] = len(all_files)
        n = 0
        t0 = time.time()
        for f in all_files:
            try:
                taille = round(f.stat().st_size/(1024*1024), 2)
                tf     = type_f(f)
                s, sa, ep = (None,None,None)
                if tf == "Vidéo": s, sa, ep = detecter_serie(f.name)
                applied = False
                for rule in load_rules():
                    src_ok = not rule.get("source","") or str(f).startswith(rule.get("source",""))
                    if src_ok and apply_rule(rule, f):
                        applied = True; break
                if not applied:
                    insert_file(f.name, tf, disque, f, taille, s, sa, ep, f.stem)
                n += 1
                scan_progress["current"] = n
                scan_progress["count"]   = n
            except: continue
        duree = round(time.time()-t0, 2)
        try:
            con = get_db()
            con.execute("INSERT INTO scan_history (date,dossier,nb_ajoutes,duree_sec) VALUES (?,?,?,?)",
                        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), str(path), n, duree))
            con.commit(); con.close()
        except: pass
    except Exception as e:
        log.error(f"Scan error: {e}")
    finally:
        scan_progress["running"] = False
        scan_progress["done"]    = True

# ── Watchers ──────────────────────────────────────────────────────
class FolderHandler(FileSystemEventHandler):
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
    def on_created(self, event):
        if event.is_directory: return
        p = Path(event.src_path)
        if p.suffix.lower() not in TOUTES_EXT: return
        time.sleep(3)
        if not p.exists() or p.stat().st_size == 0: return
        taille = round(p.stat().st_size/(1024*1024), 2)
        tf = type_f(p)
        s, sa, ep = (None, None, None)
        if tf == "Vidéo": s, sa, ep = detecter_serie(p.name)
        applied = False
        for rule in load_rules():
            src_ok = not rule.get("source","") or str(p).startswith(rule.get("source",""))
            if src_ok and apply_rule(rule, p):
                applied = True; break
        if not applied:
            insert_file(p.name, tf, str(p.anchor).rstrip("\\/"), p, taille, s, sa, ep, p.stem)
        log.info(f"Nouveau fichier : {p.name}")

def start_watcher(folder_path):
    obs = Observer()
    obs.schedule(FolderHandler(folder_path), str(folder_path), recursive=True)
    obs.daemon = True; obs.start()
    observers.append(obs)
    return obs

def load_watched_folders():
    global watched_folders
    if WATCHED_FILE.exists():
        try:
            watched_folders = json.loads(WATCHED_FILE.read_text(encoding="utf-8"))
            for f in watched_folders:
                if Path(f).exists(): start_watcher(f)
        except: pass

def save_watched_folders():
    WATCHED_FILE.write_text(json.dumps(watched_folders, ensure_ascii=False), encoding="utf-8")

def lecteurs():
    masque = ctypes.windll.kernel32.GetLogicalDrives()
    return {l for i,l in enumerate(string.ascii_uppercase) if masque&(1<<i)}

def infos_lecteur(lettre):
    types = {0:"Inconnu",1:"Sans racine",2:"Amovible",3:"Fixe",4:"Réseau",5:"CD/DVD",6:"RAM"}
    racine = f"{lettre}:\\"
    tid = ctypes.windll.kernel32.GetDriveTypeW(racine)
    lb  = ctypes.create_unicode_buffer(261)
    ctypes.windll.kernel32.GetVolumeInformationW(racine, lb, 261, None, None, None, None, 0)
    tb  = ctypes.c_ulonglong(0)
    ctypes.windll.kernel32.GetDiskFreeSpaceExW(racine, None, ctypes.byref(tb), None)
    return {"lettre":lettre,"racine":racine,"type":types.get(tid,"Inconnu"),
            "label":lb.value or f"Disque {lettre}","total_gb":round(tb.value/(1024**3),1)}

def disk_watcher_loop():
    global pending_disk
    connus = lecteurs()
    while True:
        try:
            actuels = lecteurs()
            for l in actuels - connus:
                try: pending_disk = infos_lecteur(l)
                except: pass
            connus = actuels
        except: pass
        time.sleep(3)

# ── TMDB ──────────────────────────────────────────────────────────
def fetch_tmdb_if_missing(serie, saison):
    key = get_tmdb_key()
    if not key: return
    con = get_db()
    ex = con.execute("SELECT total_ep FROM series_meta WHERE serie=? AND saison=?", (serie,saison)).fetchone()
    con.close()
    if ex and ex["total_ep"] > 0: return
    fetch_tmdb(serie, saison, key)

def fetch_tmdb(serie, saison, key=None):
    key = key or get_tmdb_key()
    if not key: return None
    try:
        r = requests.get(f"{TMDB_URL}/search/tv", params={"api_key":key,"query":serie,"language":"fr-FR"}, timeout=10)
        results = r.json().get("results",[])
        if not results:
            r = requests.get(f"{TMDB_URL}/search/tv", params={"api_key":key,"query":serie,"language":"en-US"}, timeout=10)
            results = r.json().get("results",[])
        if not results: return None
        tmdb_id = results[0]["id"]
        r2 = requests.get(f"{TMDB_URL}/tv/{tmdb_id}/season/{saison}", params={"api_key":key,"language":"fr-FR"}, timeout=10)
        total = len(r2.json().get("episodes",[]))
        if total == 0: return None
        con = get_db()
        con.execute("""INSERT OR REPLACE INTO series_meta (serie,saison,total_ep,tmdb_id,source,updated_at)
            VALUES (?,?,?,?,'tmdb',?)""", (serie,saison,total,tmdb_id,datetime.now().strftime("%Y-%m-%d %H:%M")))
        con.commit(); con.close()
        return total
    except Exception as e: log.error(f"TMDB: {e}"); return None

# ── Ollama ────────────────────────────────────────────────────────
def call_ollama_prompt(prompt: str, timeout: int = 120) -> str:
    try:
        r = requests.post(OLLAMA_URL, json={"model":MODELE,"prompt":prompt,"stream":False}, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return data.get("response", "") or ""
    except Exception as e:
        log.error(f"Ollama: {e}")
        return ""

def extract_json_object(text: str) -> dict:
    if not text:
        return {}
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        return {}
    try:
        return json.loads(m.group())
    except Exception:
        return {}

def normalize_tags_value(value) -> str:
    if isinstance(value, list):
        raw = value
    else:
        raw = re.split(r"[,;|\n]", str(value or ""))
    out, seen = [], set()
    for item in raw:
        tag = re.sub(r"\s+", " ", str(item).strip(" #.-_\t\r\n")).strip()

        if not tag:
            continue
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(tag[:32])
    return ", ".join(out[:8])

def generate_tags_from_filename(nom: str, type_fichier: str = "", serie: str = "") -> str:
    prompt = (
        "Tu aides à classer des fichiers multimédias.\n"
        "Réponds uniquement en JSON sous la forme {\"tags\":[\"tag1\",\"tag2\"]}.\n"
        "Règles :\n"
        "- déduis des tags courts et utiles uniquement à partir du nom de fichier et des métadonnées fournies\n"
        "- au maximum 6 tags\n"
        "- pas de phrase complète, pas d'explication\n"
        "- utilise le français quand c'est naturel\n\n"
        f"Nom du fichier: {nom}\n"
        f"Type: {type_fichier or 'Inconnu'}\n"
        f"Série: {serie or 'Aucune'}\n"
    )
    data = extract_json_object(call_ollama_prompt(prompt, timeout=90))
    tags = normalize_tags_value(data.get("tags", ""))
    if tags:
        return tags
    # fallback léger sans IA si jamais Ollama répond mal
    stem = Path(nom).stem
    parts = re.split(r"[^A-Za-z0-9À-ÿ]+", stem)
    banned = {"x264","x265","h264","h265","webrip","webdl","bluray","1080p","720p","2160p","pdf","jpg","jpeg","png","mp3","flac","mkv","mp4"}
    fallback = []
    for p in parts:
        if len(p) < 3:
            continue
        key = p.lower()
        if key in banned or key.isdigit():
            continue
        if key not in [x.lower() for x in fallback]:
            fallback.append(p)
        if len(fallback) >= 4:
            break
    return ", ".join(fallback)

def apply_autotag_to_row(row, overwrite: bool = False) -> dict:
    current_tags = (row["tags"] or "").strip()
    if current_tags and not overwrite:
        return {"ok": True, "skipped": True, "tags": current_tags}
    tags = generate_tags_from_filename(row["nom"] or "", row["type"] or "", row["serie"] or "")
    if not tags:
        return {"ok": False, "error": "Aucun tag généré"}
    con = get_db()
    con.execute("UPDATE fichiers SET tags=? WHERE id=?", (tags, row["id"]))
    con.commit(); con.close()
    return {"ok": True, "skipped": False, "tags": tags}

def analyse_pdf(chemin):
    try:
        import pdfplumber
        with pdfplumber.open(chemin) as pdf:
            texte = "".join(p.extract_text() or "" for p in pdf.pages[:5])[:2000]
    except: return "","",""
    prompt = (f"PDF:\n\n{texte}\n\n"
              'Réponds uniquement en JSON : {"titre":"...","resume":"2-3 phrases","tags":"tag1,tag2,tag3"}')
    try:
        response = call_ollama_prompt(prompt, timeout=120)
        m = re.search(r"\{.*\}", response, re.DOTALL)
        if m:
            d = json.loads(m.group())
            return d.get("titre",""), d.get("resume",""), d.get("tags","")
    except Exception as e: log.error(f"Ollama: {e}")
    return "","",""

# ── Export ────────────────────────────────────────────────────────
def do_export(dest):
    from openpyxl.styles import Font, PatternFill
    con  = get_db()
    rows = con.execute("SELECT * FROM fichiers ORDER BY serie,saison,episode,nom").fetchall()
    wb   = openpyxl.Workbook()
    ws   = wb.active; ws.title = "Fichiers"
    hdrs = ["Date","Nom","Type","Disque","Mo","Série","Saison","Épisode","Qualité","Titre","Tags Ollama","Tags Manuels","Note","Favori","Vu"]
    ws.append(hdrs)
    fill = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=fill
    for r in rows:
        if is_private(r["chemin"] or ""): continue
        ws.append([r["date_scan"],r["nom"],r["type"],r["disque"],r["taille_mb"],
                   r["serie"] or "",r["saison"] or "",r["episode"] or "",r["qualite"] or "",
                   r["titre"] or "",r["tags"] or "",r["tags_manuels"] or "",r["note"] or "",
                   "⭐" if r["favori"] else "","✓" if r["vu"] else ""])
    con.close(); wb.save(dest)

def mask_private(rows, unlocked: bool):
    result = []
    for d in [dict(r) for r in rows]:
        if is_private(d.get("chemin","")):
            if not unlocked:
                d["_prive"]=True; d["nom"]="[Fichier privé]"
                for k in ["titre","tags","resume","serie","qualite","tags_manuels","note"]:
                    d[k]="" if k!="serie" else None
                d["chemin"]=""
            else:
                d["_prive"]=True
        result.append(d)
    return result

# ── Lifespan (remplace @app.on_event) ────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    load_watched_folders()
    start_watcher(DOWNLOADS)
    threading.Thread(target=disk_watcher_loop, daemon=True).start()
    threading.Thread(target=backup_loop, daemon=True).start()
    log.info("Backend démarré sur http://localhost:8765")
    yield

app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ════════════════════════════════════════════════════════════════
# ROUTES
# ════════════════════════════════════════════════════════════════

@app.get("/ping")
def ping(): return {"ok": True}

@app.get("/pick-folder")
def pick_folder(default_path:str=""):
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askdirectory(initialdir=default_path or str(DOWNLOADS))
        root.destroy()
        return {"ok":True,"path":path or ""}
    except Exception as e:
        return {"ok":False,"path":"","error":str(e)}

@app.get("/scan/progress")
def get_scan_progress(): return scan_progress

class ScanReq(BaseModel): path: str

@app.post("/scan")
def post_scan(req: ScanReq):
    global scan_progress
    if scan_progress["running"]:
        return {"ok": False, "error": "Scan déjà en cours"}
    scan_progress = {"running":True,"current":0,"total":0,"folder":req.path,"done":False,"count":0}
    threading.Thread(target=scan_folder_background, args=(req.path,), daemon=True).start()
    return {"ok": True, "started": True}

@app.get("/files")
def get_files(q:str="", type:str="Tous", disk:str="Tous",
              sort:str="nom", order:str="asc", recent:int=0,
              unlocked:int=0, favoris:int=0, tag_manuel:str=""):
    con = get_db()
    conds, params = [], []
    if q:
        qq = f"%{q}%"
        conds.append("(nom LIKE ? OR serie LIKE ? OR titre LIKE ? OR tags LIKE ? OR resume LIKE ? OR tags_manuels LIKE ? OR note LIKE ?)")
        params.extend([qq,qq,qq,qq,qq,qq,qq])
    if type != "Tous": conds.append("type=?"); params.append(type)
    if disk != "Tous": conds.append("disque=?"); params.append(disk)
    if favoris: conds.append("favori=1")
    if tag_manuel: conds.append("tags_manuels LIKE ?"); params.append(f"%{tag_manuel}%")
    where = ("WHERE "+" AND ".join(conds)) if conds else ""
    sort_map = {"nom":"nom","date":"date_scan","taille":"taille_mb","type":"type","disque":"disque"}
    sort_col = sort_map.get(sort,"nom")
    sort_dir = "DESC" if order=="desc" else "ASC"
    limit = f"LIMIT {recent}" if recent>0 else "LIMIT 2000"
    order_clause = "ORDER BY date_scan DESC" if recent>0 else f"ORDER BY {sort_col} {sort_dir}"
    rows = con.execute(
        f"SELECT id,date_scan,nom,type,disque,taille_mb,serie,saison,episode,titre,resume,tags,"
        f"chemin,qualite,vu,favori,tags_manuels,note "
        f"FROM fichiers {where} {order_clause} {limit}", params
    ).fetchall()
    con.close()
    return mask_private(rows, bool(unlocked))

@app.get("/files/private")
def get_private_files(unlocked:int=0):
    pf = get_private_folders()
    if not pf: return []
    con = get_db()
    conds  = " OR ".join("chemin LIKE ?" for _ in pf)
    params = [f"{f}%" for f in pf]
    rows = con.execute(
        f"SELECT id,date_scan,nom,type,disque,taille_mb,serie,saison,episode,titre,resume,tags,"
        f"chemin,qualite,vu,favori,tags_manuels,note "
        f"FROM fichiers WHERE {conds} ORDER BY nom LIMIT 1000", params
    ).fetchall()
    con.close()
    return mask_private(rows, bool(unlocked))

@app.get("/tags-manuels")
def get_all_tags_manuels():
    con = get_db()
    rows = con.execute("SELECT tags_manuels FROM fichiers WHERE tags_manuels != '' AND tags_manuels IS NOT NULL").fetchall()
    con.close()
    all_tags = set()
    for r in rows:
        for t in (r["tags_manuels"] or "").split(","):
            t = t.strip()
            if t: all_tags.add(t)
    return sorted(all_tags)

@app.get("/series")
def get_series(q:str=""):
    con  = get_db()
    rows = con.execute("""SELECT serie,COUNT(*) nb,GROUP_CONCAT(DISTINCT disque) disques,
        MIN(saison) saison_min,MAX(saison) saison_max FROM fichiers
        WHERE serie IS NOT NULL AND serie!='' GROUP BY serie ORDER BY serie""").fetchall()
    con.close()
    data = [dict(r) for r in rows]
    if q: data = [d for d in data if q.lower() in (d["serie"]or"").lower()]
    return data

@app.get("/series/{serie_name}/detail")
def get_serie_detail(serie_name:str):
    con = get_db()
    rows = con.execute("""SELECT saison,episode,nom,disque,taille_mb,chemin,qualite,vu
        FROM fichiers WHERE serie=? AND type='Vidéo' ORDER BY saison,episode""", (serie_name,)).fetchall()
    metas = con.execute("SELECT saison,total_ep,source,tmdb_id FROM series_meta WHERE serie=?", (serie_name,)).fetchall()
    con.close()
    meta_map = {m["saison"]:dict(m) for m in metas}
    saisons = {}
    for r in rows:
        s = r["saison"] or 0
        if s not in saisons: saisons[s] = []
        saisons[s].append({"episode":r["episode"],"nom":r["nom"],"disque":r["disque"],
                           "taille_mb":r["taille_mb"],"chemin":r["chemin"],
                           "qualite":r["qualite"] or "","vu":r["vu"] or 0})
    result = []
    for s in sorted(saisons.keys()):
        eps=saisons[s]; meta=meta_map.get(s,{})
        result.append({"saison":s,"episodes":eps,"count":len(eps),
                       "total_ep":meta.get("total_ep",0),"source":meta.get("source",""),"tmdb_id":meta.get("tmdb_id")})
    return {"serie":serie_name,"saisons":result}

@app.get("/stats")
def get_stats():
    con = get_db()
    par_type   = con.execute("SELECT type,COUNT(*) count,ROUND(SUM(taille_mb),0) size_mb FROM fichiers GROUP BY type ORDER BY count DESC").fetchall()
    par_disk   = con.execute("SELECT disque,COUNT(*) count FROM fichiers GROUP BY disque ORDER BY count DESC").fetchall()
    nb_favoris = con.execute("SELECT COUNT(*) FROM fichiers WHERE favori=1").fetchone()[0]
    con.close()
    return {"par_type":[dict(r) for r in par_type],"par_disk":[dict(r) for r in par_disk],"nb_favoris":nb_favoris}

@app.get("/disks")
def get_disks():
    con  = get_db()
    rows = con.execute("SELECT DISTINCT disque FROM fichiers ORDER BY disque").fetchall()
    con.close()
    return [r["disque"] for r in rows]

@app.get("/duplicates")
def get_duplicates():
    con  = get_db()
    rows = con.execute("""SELECT nom,type,COUNT(*) cnt,GROUP_CONCAT(disque) disques,
        GROUP_CONCAT(chemin,'|||') chemins,ROUND(SUM(taille_mb),1) total_mb,taille_mb taille_unit
        FROM fichiers GROUP BY nom,taille_mb HAVING cnt>1 ORDER BY total_mb DESC LIMIT 500""").fetchall()
    con.close()
    result = []
    for r in rows:
        chemins = r["chemins"].split("|||")
        if is_private(chemins[0]): continue
        result.append({"nom":r["nom"],"type":r["type"],"count":r["cnt"],
                       "disques":r["disques"],"chemins":chemins,
                       "total_mb":r["total_mb"],"taille_unit":r["taille_unit"]})
    return result

@app.get("/missing-episodes")
def get_missing_episodes():
    con   = get_db()
    metas = con.execute("SELECT serie,saison,total_ep FROM series_meta WHERE total_ep>0").fetchall()
    result = []
    for meta in metas:
        serie,saison,total = meta["serie"],meta["saison"],meta["total_ep"]
        present = {r["episode"] for r in con.execute(
            "SELECT episode FROM fichiers WHERE serie=? AND saison=? AND type='Vidéo'", (serie,saison)).fetchall()}
        missing = [i for i in range(1,total+1) if i not in present]
        if missing: result.append({"serie":serie,"saison":saison,"manquants":missing,"count":len(missing)})
    con.close()
    result.sort(key=lambda x:(x["serie"],x["saison"]))
    return result

@app.get("/scan-history")
def get_scan_history():
    con  = get_db()
    rows = con.execute("SELECT * FROM scan_history ORDER BY date DESC LIMIT 50").fetchall()
    con.close()
    return [dict(r) for r in rows]

@app.get("/corrupted")
def get_corrupted():
    con  = get_db()
    rows = con.execute("SELECT id,nom,type,disque,chemin,taille_mb FROM fichiers ORDER BY nom").fetchall()
    con.close()
    result = []
    for r in rows:
        raison = None
        try:
            p = Path(r["chemin"])
            if not p.exists():                raison = "Fichier introuvable"
            elif p.stat().st_size == 0:       raison = "Fichier vide (0 octets)"
            elif (r["taille_mb"] or 0) == 0:  raison = "Taille 0 MB"
        except: raison = "Erreur d'accès"
        if raison:
            result.append({"id":r["id"],"nom":r["nom"],"type":r["type"],
                           "disque":r["disque"],"chemin":r["chemin"],
                           "taille_mb":r["taille_mb"],"raison":raison})
    return result

@app.delete("/corrupted/{file_id}")
def remove_corrupted(file_id:int):
    con = get_db()
    con.execute("DELETE FROM fichiers WHERE id=?", (file_id,))
    con.commit(); con.close()
    return {"ok":True}

class CompareReq(BaseModel):
    dossier_a: str
    dossier_b: str

@app.post("/compare-folders")
def compare_folders(req:CompareReq):
    def scan(path):
        p = Path(path)
        if not p.exists(): return {}
        result = {}
        for f in p.rglob("*"):
            try:
                if f.is_file():
                    result[str(f.relative_to(p))] = round(f.stat().st_size/(1024*1024),2)
            except: continue
        return result
    fa,fb = scan(req.dossier_a), scan(req.dossier_b)
    ka,kb = set(fa.keys()), set(fb.keys())
    return {
        "dossier_a":req.dossier_a,"dossier_b":req.dossier_b,
        "seulement_a":[{"nom":k,"taille_mb":fa[k]} for k in sorted(ka-kb)],
        "seulement_b":[{"nom":k,"taille_mb":fb[k]} for k in sorted(kb-ka)],
        "communs":[{"nom":k,"taille_a":fa[k],"taille_b":fb[k],"diff_taille":abs(fa[k]-fb[k])>0.01} for k in sorted(ka&kb)],
        "total_a":len(fa),"total_b":len(fb),
        "nb_communs":len(ka&kb),"nb_seulement_a":len(ka-kb),"nb_seulement_b":len(kb-ka),
    }

@app.get("/stats/evolution")
def get_evolution():
    con  = get_db()
    rows = con.execute("""
        SELECT strftime('%Y-W%W', date_scan) as semaine, COUNT(*) as nb,
               ROUND(SUM(taille_mb),0) as size_mb
        FROM fichiers WHERE date_scan >= date('now','-84 days')
        GROUP BY semaine ORDER BY semaine ASC""").fetchall()
    con.close()
    return [dict(r) for r in rows]

class ExportReq(BaseModel): path: str

@app.post("/export")
def post_export(req:ExportReq):
    try: do_export(req.path); return {"ok":True}
    except Exception as e: return {"ok":False,"error":str(e)}

class AnalyseReq(BaseModel): chemin: str

@app.post("/analyse")
def post_analyse(req:AnalyseReq):
    titre,resume,tags = analyse_pdf(req.chemin)
    if not resume: return {"ok":False}
    con = get_db()
    con.execute("UPDATE fichiers SET titre=?,resume=?,tags=?,analyse=1 WHERE chemin=?", (titre,resume,tags,req.chemin))
    con.commit(); con.close()
    return {"ok":True,"titre":titre,"resume":resume,"tags":tags}

@app.post("/analyse-batch")
def post_analyse_batch():
    con  = get_db()
    rows = con.execute("SELECT id,chemin,nom FROM fichiers WHERE type='PDF' AND analyse=0 LIMIT 50").fetchall()
    con.close()
    total = len(rows)
    if total == 0: return {"ok":True,"count":0}
    def run():
        for r in rows:
            if is_private(r["chemin"]): continue
            titre,resume,tags = analyse_pdf(r["chemin"])
            if resume:
                c = get_db()
                c.execute("UPDATE fichiers SET titre=?,resume=?,tags=?,analyse=1 WHERE id=?", (titre,resume,tags,r["id"]))
                c.commit(); c.close()
    threading.Thread(target=run, daemon=True).start()
    return {"ok":True,"count":total}

@app.get("/analyse-batch-status")
def get_batch_status():
    con = get_db()
    total    = con.execute("SELECT COUNT(*) FROM fichiers WHERE type='PDF'").fetchone()[0]
    analysed = con.execute("SELECT COUNT(*) FROM fichiers WHERE type='PDF' AND analyse=1").fetchone()[0]
    con.close()
    return {"total":total,"analysed":analysed,"remaining":total-analysed}

class VuReq(BaseModel): chemin:str; vu:int

@app.post("/vu")
def set_vu(req:VuReq):
    con = get_db()
    con.execute("UPDATE fichiers SET vu=? WHERE chemin=?", (req.vu,req.chemin))
    con.commit(); con.close()
    return {"ok":True}

class FavoriReq(BaseModel): chemin:str; favori:int

@app.post("/favori")
def set_favori(req:FavoriReq):
    con = get_db()
    con.execute("UPDATE fichiers SET favori=? WHERE chemin=?", (req.favori,req.chemin))
    con.commit(); con.close()
    return {"ok":True,"favori":req.favori}

class TagsManuelReq(BaseModel): chemin:str; tags_manuels:str

@app.post("/tags-manuels")
def set_tags_manuels(req:TagsManuelReq):
    con = get_db()
    con.execute("UPDATE fichiers SET tags_manuels=? WHERE chemin=?", (req.tags_manuels.strip(),req.chemin))
    con.commit(); con.close()
    return {"ok":True}

class NoteReq(BaseModel): chemin:str; note:str

@app.post("/note")
def set_note(req:NoteReq):
    con = get_db()
    con.execute("UPDATE fichiers SET note=? WHERE chemin=?", (req.note,req.chemin))
    con.commit(); con.close()
    return {"ok":True}

class FixSerieReq(BaseModel):
    chemin:str; serie:str; saison:int; episode:int; nom_affichage:str=""

@app.post("/fix-serie")
def fix_serie(req:FixSerieReq):
    con = get_db()
    titre = req.nom_affichage or req.serie
    con.execute("UPDATE fichiers SET serie=?,saison=?,episode=?,titre=? WHERE chemin=?",
                (req.serie,req.saison,req.episode,titre,req.chemin))
    con.commit(); con.close()
    return {"ok":True}

class OuvrirReq(BaseModel): chemin:str

@app.post("/ouvrir")
def ouvrir_fichier(req:OuvrirReq):
    import os
    try: os.startfile(req.chemin); return {"ok":True}
    except Exception as e: return {"ok":False,"error":str(e)}

@app.get("/new-disk")
def new_disk():
    global pending_disk
    if pending_disk: d=pending_disk; return {"disk":d}
    return {"disk":None}

@app.post("/dismiss-disk")
def dismiss_disk():
    global pending_disk; pending_disk=None; return {"ok":True}

@app.get("/models")
def get_models():
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        return {"models":[m["name"] for m in r.json().get("models",[])],"current":MODELE}
    except: return {"models":[],"current":MODELE}

class ModelReq(BaseModel): model:str

@app.post("/model")
def set_model(req:ModelReq):
    global MODELE; MODELE=req.model; return {"ok":True,"model":MODELE}

@app.get("/watched-folders")
def get_watched_folders(): return {"folders":[str(DOWNLOADS)]+watched_folders}

class FolderReq(BaseModel): path:str

@app.post("/watched-folders")
def add_watched_folder(req:FolderReq):
    global watched_folders
    path=req.path.strip()
    if path in watched_folders or path==str(DOWNLOADS): return {"ok":False,"error":"Dossier déjà surveillé"}
    if not Path(path).exists(): return {"ok":False,"error":"Dossier introuvable"}
    watched_folders.append(path); save_watched_folders(); start_watcher(path)
    scan_progress_save = dict(scan_progress)
    n_ref = [0]
    def count_cb(n): n_ref[0]=n
    t = threading.Thread(target=lambda: n_ref.__setitem__(0, len([f for f in Path(path).rglob("*") if f.is_file() and f.suffix.lower() in TOUTES_EXT])), daemon=True)
    t.start(); t.join(timeout=5)
    threading.Thread(target=scan_folder_background, args=(path,), daemon=True).start()
    return {"ok":True,"count":0,"folders":[str(DOWNLOADS)]+watched_folders}

@app.delete("/watched-folders")
def remove_watched_folder(req:FolderReq):
    global watched_folders
    if req.path in watched_folders: watched_folders.remove(req.path); save_watched_folders()
    return {"ok":True,"folders":[str(DOWNLOADS)]+watched_folders}

@app.get("/config")
def get_config():
    c=load_config()
    return {"tmdb_key":c.get("tmdb_key",""),"tmdb_key_set":bool(c.get("tmdb_key","")),
            "pin_set":bool(c.get("pin_hash","")),"private_timeout":c.get("private_timeout",5),
            "backup_enabled":c.get("backup_enabled",True),"backup_keep":c.get("backup_keep",7),
            "backup_dir":c.get("backup_dir",str(BACKUP_DIR))}

class ConfigReq(BaseModel):
    tmdb_key:str=""; private_timeout:int=5
    backup_enabled:bool=True; backup_keep:int=7; backup_dir:str=""

@app.post("/config")
def save_config_route(req:ConfigReq):
    c=load_config()
    c["tmdb_key"]=req.tmdb_key.strip(); c["private_timeout"]=req.private_timeout
    c["backup_enabled"]=req.backup_enabled; c["backup_keep"]=req.backup_keep
    c["backup_dir"]=req.backup_dir or str(BACKUP_DIR)
    save_config(c); return {"ok":True}

@app.post("/backup/now")
def backup_now():
    try:
        dest=do_backup()
        if dest: return {"ok":True,"path":dest}
        return {"ok":False,"error":"Backup désactivé ou DB introuvable"}
    except Exception as e: return {"ok":False,"error":str(e)}

@app.get("/backup/list")
def backup_list():
    c=load_config(); backup_dir=Path(c.get("backup_dir",str(BACKUP_DIR)))
    if not backup_dir.exists(): return {"backups":[]}
    backups=sorted(backup_dir.glob("media_indexer_*.db"),key=lambda p:p.stat().st_mtime,reverse=True)
    return {"backups":[{"nom":b.name,"taille_mb":round(b.stat().st_size/(1024*1024),2),
                        "date":datetime.fromtimestamp(b.stat().st_mtime).strftime("%Y-%m-%d %H:%M")} for b in backups]}

class RestoreReq(BaseModel): nom:str

@app.post("/backup/restore")
def backup_restore(req:RestoreReq):
    c=load_config(); backup_dir=Path(c.get("backup_dir",str(BACKUP_DIR)))
    src=backup_dir/req.nom
    if not src.exists(): return {"ok":False,"error":"Fichier introuvable"}
    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(DB_FILE, backup_dir/f"before_restore_{ts}.db")
    shutil.copy2(src, DB_FILE)
    return {"ok":True}

@app.get("/pin/status")
def pin_status(): return {"configured":bool(get_pin_hash())}

class PinSetReq(BaseModel): pin:str

@app.post("/pin/set")
def set_pin(req:PinSetReq):
    if len(req.pin)<4: return {"ok":False,"error":"PIN minimum 4 caractères"}
    c=load_config(); c["pin_hash"]=hashlib.sha256(req.pin.encode()).hexdigest()
    save_config(c); return {"ok":True}

class PinVerifyReq(BaseModel): pin:str

@app.post("/pin/verify")
def verify_pin_route(req:PinVerifyReq):
    if verify_pin(req.pin): return {"ok":True}
    return {"ok":False,"error":"PIN incorrect"}

@app.post("/pin/delete")
def delete_pin():
    c=load_config(); c.pop("pin_hash",None); save_config(c); return {"ok":True}

@app.get("/private-folders")
def get_private_folders_route(): return {"folders":get_private_folders()}

class PrivateFolderReq(BaseModel): path:str

@app.post("/private-folders")
def add_private_folder(req:PrivateFolderReq):
    c=load_config(); folders=c.get("private_folders",[])
    if req.path not in folders: folders.append(req.path)
    c["private_folders"]=folders; save_config(c); return {"ok":True,"folders":folders}

@app.delete("/private-folders")
def remove_private_folder(req:PrivateFolderReq):
    c=load_config(); folders=c.get("private_folders",[])
    if req.path in folders: folders.remove(req.path)
    c["private_folders"]=folders; save_config(c); return {"ok":True,"folders":folders}

class TmdbFetchReq(BaseModel): serie:str; saison:int

@app.post("/tmdb-fetch")
def tmdb_fetch(req:TmdbFetchReq):
    key=get_tmdb_key()
    if not key: return {"ok":False,"error":"Clé TMDB non configurée"}
    total=fetch_tmdb(req.serie,req.saison,key)
    if total: return {"ok":True,"total":total}
    return {"ok":False,"error":f"'{req.serie}' introuvable sur TMDB"}

class SerieMetaReq(BaseModel): serie:str; saison:int; total_ep:int

@app.post("/series-meta")
def set_serie_meta(req:SerieMetaReq):
    con=get_db()
    con.execute("""INSERT OR REPLACE INTO series_meta (serie,saison,total_ep,source,updated_at)
        VALUES (?,?,?,'manuel',?)""", (req.serie,req.saison,req.total_ep,datetime.now().strftime("%Y-%m-%d %H:%M")))
    con.commit(); con.close(); return {"ok":True}

@app.get("/rules")
def get_rules(): return load_rules()

class RuleModel(BaseModel):
    id:str=""; nom:str; source:str; condition_type:str
    condition_val:str; action:str; destination:str; actif:bool=True

@app.post("/rules")
def add_rule(req:RuleModel):
    rules=load_rules(); rule=req.dict(); rule["id"]=str(uuid.uuid4())[:8]
    rules.append(rule); save_rules(rules); return {"ok":True,"rules":rules}

@app.delete("/rules/{rule_id}")
def delete_rule(rule_id:str):
    rules=[r for r in load_rules() if r["id"]!=rule_id]
    save_rules(rules); return {"ok":True,"rules":rules}

class AutoTagReq(BaseModel):
    chemin:str
    nom:str=""
    type:str=""
    serie:str=""

@app.post("/autotag")
def post_autotag(req: AutoTagReq):
    con = get_db()
    row = con.execute("SELECT id,nom,type,serie,tags FROM fichiers WHERE chemin=?", (req.chemin,)).fetchone()
    con.close()
    if not row:
        return {"ok": False, "error": "Fichier introuvable"}
    result = apply_autotag_to_row(row, overwrite=True)
    if not result.get("ok"):
        return {"ok": False, "error": result.get("error", "Erreur auto-tag") }
    return {"ok": True, "tags": result.get("tags", "")}

class AutoTagFolderReq(BaseModel):
    folder:str
    limit:int=200
    overwrite:bool=False
    only_untagged:bool=True

@app.post("/autotag/folder/start")
def start_autotag_folder(req: AutoTagFolderReq):
    global autotag_progress, autotag_stop_flag
    if autotag_progress.get("running"):
        return {"ok": False, "error": "Un auto-tag dossier est déjà en cours"}
    folder = (req.folder or "").strip()
    if not folder:
        return {"ok": False, "error": "Dossier manquant"}
    limit = max(1, min(int(req.limit or 200), 1000))
    like = folder.rstrip('\\/') + '%'
    conds = ["chemin LIKE ?"]
    params = [like]
    if req.only_untagged and not req.overwrite:
        conds.append("(tags IS NULL OR TRIM(tags)='')")
    con = get_db()
    rows = con.execute(f"SELECT id,nom,type,serie,tags,chemin FROM fichiers WHERE {' AND '.join(conds)} ORDER BY nom LIMIT ?", (*params, limit)).fetchall()
    con.close()
    total = len(rows)
    autotag_stop_flag = False
    autotag_progress = {"running": total > 0, "done": total == 0, "folder": folder, "current": "", "total": total, "processed": 0, "tagged": 0, "skipped": 0, "errors": 0, "last_error": "", "overwrite": bool(req.overwrite), "only_untagged": bool(req.only_untagged)}
    if total == 0:
        return {"ok": True, "started": False, "count": 0}

    def worker(rows_snapshot):
        global autotag_progress, autotag_stop_flag
        for row in rows_snapshot:
            if autotag_stop_flag:
                break
            autotag_progress["current"] = row["nom"]
            try:
                result = apply_autotag_to_row(row, overwrite=req.overwrite)
                if result.get("ok"):
                    if result.get("skipped"):
                        autotag_progress["skipped"] += 1
                    else:
                        autotag_progress["tagged"] += 1
                else:
                    autotag_progress["errors"] += 1
                    autotag_progress["last_error"] = result.get("error", "Erreur")
            except Exception as e:
                autotag_progress["errors"] += 1
                autotag_progress["last_error"] = str(e)
            finally:
                autotag_progress["processed"] += 1
        autotag_progress["running"] = False
        autotag_progress["done"] = True
        autotag_progress["current"] = ""

    threading.Thread(target=worker, args=(rows,), daemon=True).start()
    return {"ok": True, "started": True, "count": total}

@app.get("/autotag/folder/status")
def get_autotag_folder_status():
    return autotag_progress

@app.post("/autotag/folder/stop")
def stop_autotag_folder():
    global autotag_stop_flag
    autotag_stop_flag = True
    return {"ok": True}


@app.get("/regroup/suggestions")
def get_regroup_suggestions():
    con = get_db()
    suggestions = []
    # Séries sur plusieurs disques
    rows = con.execute("""SELECT serie,COUNT(DISTINCT disque) nb_disques,
        GROUP_CONCAT(DISTINCT disque) disques,COUNT(*) nb_fichiers
        FROM fichiers WHERE serie IS NOT NULL AND serie!='' AND type='Vidéo'
        GROUP BY serie HAVING nb_disques>1 ORDER BY nb_fichiers DESC""").fetchall()
    for r in rows:
        eps = con.execute("SELECT chemin,saison,episode,nom,disque FROM fichiers WHERE serie=? AND type='Vidéo' ORDER BY saison,episode", (r["serie"],)).fetchall()
        suggestions.append({"type":"serie_multi_disques","titre":r["serie"],
            "description":f"Épisodes sur {r['nb_disques']} disques ({r['disques']})",
            "nb_fichiers":r["nb_fichiers"],"disques":r["disques"].split(","),
            "fichiers":[{"chemin":e["chemin"],"saison":e["saison"],"episode":e["episode"],"nom":e["nom"],"disque":e["disque"]} for e in eps]})
    # Séries avec dossiers différents (même disque)
    series_1disk = con.execute("""SELECT serie,GROUP_CONCAT(DISTINCT disque) disques,COUNT(*) nb_fichiers
        FROM fichiers WHERE serie IS NOT NULL AND serie!='' AND type='Vidéo'
        GROUP BY serie HAVING COUNT(DISTINCT disque)=1""").fetchall()
    for r in series_1disk:
        eps = con.execute("SELECT chemin,saison,episode,nom,disque FROM fichiers WHERE serie=? AND type='Vidéo'", (r["serie"],)).fetchall()
        dossiers = set(str(Path(e["chemin"]).parent) for e in eps)
        if len(dossiers) > 1:
            suggestions.append({"type":"serie_multi_dossiers","titre":r["serie"],
                "description":f"Saisons dans {len(dossiers)} dossiers différents",
                "nb_fichiers":r["nb_fichiers"],"disques":[r["disques"]],
                "fichiers":[{"chemin":e["chemin"],"saison":e["saison"],"episode":e["episode"],"nom":e["nom"],"disque":e["disque"]} for e in eps],
                "dossiers":list(dossiers)})
    # Types sur plusieurs disques
    types_rows = con.execute("""SELECT type,COUNT(DISTINCT disque) nb_disques,
        GROUP_CONCAT(DISTINCT disque) disques,COUNT(*) nb_fichiers,ROUND(SUM(taille_mb),0) taille_total
        FROM fichiers WHERE type IN ('PDF','Musique','Image')
        GROUP BY type HAVING nb_disques>1 ORDER BY taille_total DESC""").fetchall()
    for r in types_rows:
        suggestions.append({"type":"type_multi_disques","titre":f"Fichiers {r['type']}",
            "description":f"{r['nb_fichiers']} fichiers sur {r['nb_disques']} disques ({r['disques']})",
            "nb_fichiers":r["nb_fichiers"],"disques":r["disques"].split(","),
            "taille_mb":r["taille_total"],"file_type":r["type"],"fichiers":[]})
    con.close()
    return suggestions

class RegroupReq(BaseModel):
    fichiers:list; destination:str; action:str="deplacer"; organiser_saisons:bool=True

@app.post("/regroup/preview")
def preview_regroup(req:RegroupReq):
    dest_base=Path(req.destination); preview=[]
    for chemin in req.fichiers:
        src=Path(chemin); dest_dir=dest_base
        if req.organiser_saisons:
            _,saison,_ = detecter_serie(src.name)
            if saison: dest_dir=dest_base/f"Saison {saison:02d}"
        dest=dest_dir/src.name
        preview.append({"nom":src.name,"source":chemin,"destination":str(dest),
                        "existe_deja":dest.exists() and dest!=src,"source_existe":src.exists()})
    return {"preview":preview,"destination_base":str(dest_base)}

@app.post("/regroup/execute")
def execute_regroup(req:RegroupReq):
    dest_base=Path(req.destination); dest_base.mkdir(parents=True,exist_ok=True)
    results={"ok":[],"errors":[],"skipped":[]}
    for chemin in req.fichiers:
        src=Path(chemin)
        if not src.exists(): results["skipped"].append({"chemin":chemin,"raison":"Introuvable"}); continue
        dest_dir=dest_base
        if req.organiser_saisons:
            _,saison,_=detecter_serie(src.name)
            if saison: dest_dir=dest_base/f"Saison {saison:02d}"; dest_dir.mkdir(parents=True,exist_ok=True)
        dest=dest_dir/src.name
        if dest.exists() and dest!=src: results["skipped"].append({"chemin":chemin,"raison":"Déjà présent"}); continue
        try:
            if req.action=="deplacer": shutil.move(str(src),str(dest))
            else: shutil.copy2(str(src),str(dest))
            con=get_db(); con.execute("UPDATE fichiers SET chemin=?,disque=? WHERE chemin=?",
                (str(dest),str(dest.anchor).rstrip("\\/"),chemin)); con.commit(); con.close()
            results["ok"].append(str(dest))
        except Exception as e: results["errors"].append({"chemin":chemin,"raison":str(e)})
    return {"ok":True,"déplacés":len(results["ok"]),"erreurs":len(results["errors"]),"ignorés":len(results["skipped"]),"details":results}

@app.post("/shutdown")
def shutdown_backend():
    def _stop():
        time.sleep(0.2)
        os._exit(0)
    threading.Thread(target=_stop, daemon=True).start()
    return {"ok": True, "message": "Arrêt du backend demandé"}

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8765, log_level="warning")