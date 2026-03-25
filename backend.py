"""
backend.py — API locale FastAPI pour Media Indexer
Tourne sur http://localhost:8765
Sert de pont entre l'UI web (Tauri) et la logique Python (SQLite, watchdog, Ollama)

Prérequis supplémentaires :
    pip install fastapi uvicorn

Lancer manuellement (pour tester) :
    python backend.py
"""

import re
import json
import time
import ctypes
import string
import logging
import sqlite3
import threading
import requests
import uvicorn
import openpyxl

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from pathlib import Path
from datetime import datetime
from typing import Optional
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ─── Config ───────────────────────────────────────────────────────────────────

DOWNLOADS_FOLDER    = Path.home() / "Downloads"
DB_FILE             = Path.home() / "Documents" / "media_indexer.db"
LOG_FILE            = Path.home() / "Documents" / "media_indexer.log"
OLLAMA_URL          = "http://localhost:11434/api/generate"
MODELE_TEXTE        = "llama3"
DISK_CHECK_INTERVAL = 3
HOST                = "127.0.0.1"
PORT                = 8765

EXT_VIDEO  = {".mp4",".mkv",".avi",".mov",".wmv",".m4v",".ts",".flv",".webm"}
EXT_PDF    = {".pdf"}
EXT_IMAGE  = {".jpg",".jpeg",".png",".gif",".webp",".bmp",".tiff",".heic"}
EXT_AUDIO  = {".mp3",".flac",".aac",".wav",".ogg",".m4a",".wma",".opus"}
EXT_AUTRES = {".txt",".md",".csv",".docx",".xlsx",".zip",".rar",".7z"}
TOUTES_EXT = EXT_VIDEO | EXT_PDF | EXT_IMAGE | EXT_AUDIO | EXT_AUTRES

DOSSIERS_IGNORES = {
    "windows","program files","program files (x86)","programdata",
    "appdata","system volume information","$recycle.bin","recovery",
    ".git","__pycache__","node_modules"
}

# ─── Logging ──────────────────────────────────────────────────────────────────

LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─── État global partagé ──────────────────────────────────────────────────────

class State:
    status          = "Prêt."
    nouveau_fichier = False
    nouveau_disque  = None   # dict infos ou None
    scan_running    = False

state = State()

# ─── Détection série ──────────────────────────────────────────────────────────

RE_EPISODE = [
    re.compile(r"[Ss](\d{1,2})[Ee](\d{1,3})"),
    re.compile(r"(\d{1,2})x(\d{1,3})"),
    re.compile(r"[Ss]eason\s*(\d+).*[Ee]p?\s*(\d+)", re.I),
    re.compile(r"[Ee]pisode\s*(\d+)", re.I),
    re.compile(r"[Ee](\d{2,3})\b"),
]

def detecter_serie(nom):
    stem  = Path(nom).stem
    clean = re.sub(r"[._]", " ", stem)
    saison, episode, match_pos = None, None, len(clean)
    for pattern in RE_EPISODE:
        m = pattern.search(clean)
        if m:
            groups = m.groups()
            saison  = groups[0]
            episode = groups[1] if len(groups) >= 2 else groups[0]
            if len(groups) == 1: saison = "1"
            match_pos = min(match_pos, m.start())
            break
    if saison is not None:
        serie_raw = clean[:match_pos].strip()
        serie = re.sub(r"[\[\(].*","",serie_raw).strip().title()
        if len(serie) < 2: serie = stem[:20].title()
        return serie, int(saison), int(episode)
    return None, None, None

def type_fichier(chemin):
    ext = Path(chemin).suffix.lower()
    if ext in EXT_VIDEO:  return "Vidéo"
    if ext in EXT_PDF:    return "PDF"
    if ext in EXT_IMAGE:  return "Image"
    if ext in EXT_AUDIO:  return "Musique"
    return "Autre"

# ─── Base de données ──────────────────────────────────────────────────────────

def get_con():
    return sqlite3.connect(DB_FILE)

def init_db():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    con = get_con()
    con.executescript("""
        CREATE TABLE IF NOT EXISTS fichiers (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            date_scan TEXT,
            nom       TEXT,
            type      TEXT,
            disque    TEXT,
            chemin    TEXT UNIQUE,
            taille_mb REAL,
            serie     TEXT,
            saison    INTEGER,
            episode   INTEGER,
            titre     TEXT,
            resume    TEXT,
            tags      TEXT,
            analyse   INTEGER DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_type   ON fichiers(type);
        CREATE INDEX IF NOT EXISTS idx_serie  ON fichiers(serie);
        CREATE INDEX IF NOT EXISTS idx_chemin ON fichiers(chemin);
    """)
    con.commit(); con.close()

def inserer_fichier(nom, type_f, disque, chemin, taille_mb,
                    serie=None, saison=None, episode=None,
                    titre="", resume="", tags=""):
    try:
        con = get_con()
        con.execute("""
            INSERT OR IGNORE INTO fichiers
            (date_scan,nom,type,disque,chemin,taille_mb,serie,saison,episode,titre,resume,tags)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (datetime.now().strftime("%Y-%m-%d %H:%M"), nom, type_f, disque,
              str(chemin), taille_mb, serie, saison, episode, titre, resume, tags))
        con.commit(); con.close()
    except Exception as e:
        log.error(f"DB insert: {e}")

def mettre_a_jour_analyse(chemin, titre, resume, tags):
    con = get_con()
    con.execute(
        "UPDATE fichiers SET titre=?,resume=?,tags=?,analyse=1 WHERE chemin=?",
        (titre, resume, tags, str(chemin))
    )
    con.commit(); con.close()

# ─── Scan ─────────────────────────────────────────────────────────────────────

def scanner_dossier(dossier):
    dossier = Path(dossier)
    disque  = str(dossier.anchor).rstrip("\\/")
    compte  = 0
    state.scan_running = True
    for chemin in dossier.rglob("*"):
        try:
            if any(p.lower() in DOSSIERS_IGNORES for p in chemin.parts): continue
            if not chemin.is_file(): continue
            if chemin.suffix.lower() not in TOUTES_EXT: continue
            taille_mb = round(chemin.stat().st_size / (1024*1024), 2)
            type_f    = type_fichier(chemin)
            serie, saison, episode = None, None, None
            if type_f == "Vidéo":
                serie, saison, episode = detecter_serie(chemin.name)
            inserer_fichier(chemin.name, type_f, disque, chemin, taille_mb,
                            serie, saison, episode, chemin.stem)
            compte += 1
            if compte % 100 == 0:
                state.status = f"Scan… {compte} fichiers ({chemin.parent.name})"
        except (PermissionError, OSError):
            continue
    state.status = f"Scan terminé — {compte} fichiers indexés."
    state.scan_running = False
    state.nouveau_fichier = True
    return compte

# ─── Ollama ───────────────────────────────────────────────────────────────────

def analyser_pdf_ollama(chemin):
    try:
        import pdfplumber
        with pdfplumber.open(chemin) as pdf:
            texte = "".join(p.extract_text() or "" for p in pdf.pages[:5])[:2000]
    except Exception:
        return "", "Lecture impossible.", ""
    prompt = (
        f"Document PDF :\n\n{texte}\n\n"
        'Réponds uniquement en JSON : {"titre":"...","resume":"résumé 2-3 phrases","tags":"tag1,tag2,tag3"}'
    )
    try:
        resp = requests.post(OLLAMA_URL,
            json={"model": MODELE_TEXTE, "prompt": prompt, "stream": False}, timeout=120)
        m = re.search(r"\{.*\}", resp.json().get("response",""), re.DOTALL)
        if m:
            d = json.loads(m.group())
            return d.get("titre",""), d.get("resume",""), d.get("tags","")
    except Exception as e:
        log.error(f"Ollama: {e}")
    return "", "", ""

# ─── Export Excel ─────────────────────────────────────────────────────────────

def exporter_excel(chemin_sortie, type_filtre="Tous", disque_filtre="Tous"):
    from openpyxl.styles import Font, PatternFill, Alignment
    con   = get_con()
    conds, params = [], []
    if type_filtre != "Tous": conds.append("type=?"); params.append(type_filtre)
    if disque_filtre != "Tous": conds.append("disque=?"); params.append(disque_filtre)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows  = con.execute(
        f"SELECT date_scan,nom,type,disque,taille_mb,serie,saison,episode,titre,tags FROM fichiers {where} ORDER BY serie,saison,episode,nom",
        params
    ).fetchall()
    series = con.execute("""
        SELECT serie,COUNT(*),GROUP_CONCAT(DISTINCT disque),MIN(saison),MAX(saison)
        FROM fichiers WHERE serie IS NOT NULL AND serie!='' GROUP BY serie ORDER BY serie
    """).fetchall()
    total    = con.execute("SELECT COUNT(*) FROM fichiers").fetchone()[0]
    par_type = con.execute("SELECT type,COUNT(*),ROUND(SUM(taille_mb),0) FROM fichiers GROUP BY type ORDER BY COUNT(*) DESC").fetchall()
    disques  = con.execute("SELECT disque,COUNT(*) FROM fichiers GROUP BY disque ORDER BY COUNT(*) DESC").fetchall()
    con.close()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Tous les fichiers"
    entetes = ["Date scan","Nom","Type","Disque","Taille (MB)","Série","Saison","Épisode","Titre","Tags"]
    ws.append(entetes)
    fill_h = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill_h
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col, w in zip("ABCDEFGHIJ", [18,35,8,8,10,25,7,7,30,25]):
        ws.column_dimensions[col].width = w

    if series:
        ws2 = wb.create_sheet("Séries")
        ws2.append(["Série","Nb épisodes","Disques","Saison min","Saison max"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7030A0")
        for s in series: ws2.append(list(s))

    ws3 = wb.create_sheet("Statistiques")
    ws3.append(["Type","Nombre","Taille totale (MB)"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="375623")
    for t,n,sz in par_type: ws3.append([t,n,sz])
    ws3.append([]); ws3.append(["Disque","Nombre"])
    for d,n in disques: ws3.append([d,n])
    wb.save(chemin_sortie)

# ─── Détection disques Windows ────────────────────────────────────────────────

def lecteurs_presents():
    masque = ctypes.windll.kernel32.GetLogicalDrives()
    return {l for i,l in enumerate(string.ascii_uppercase) if masque & (1<<i)}

def infos_lecteur(lettre):
    types = {0:"Inconnu",1:"Sans racine",2:"Amovible",3:"Fixe",4:"Réseau",5:"CD/DVD",6:"RAM"}
    racine    = f"{lettre}:\\"
    type_id   = ctypes.windll.kernel32.GetDriveTypeW(racine)
    label_buf = ctypes.create_unicode_buffer(261)
    ctypes.windll.kernel32.GetVolumeInformationW(racine, label_buf, 261, None, None, None, None, 0)
    label       = label_buf.value or f"Disque {lettre}"
    total_bytes = ctypes.c_ulonglong(0)
    ctypes.windll.kernel32.GetDiskFreeSpaceExW(racine, None, ctypes.byref(total_bytes), None)
    total_gb = round(total_bytes.value / (1024**3), 1)
    return {"lettre":lettre,"racine":racine,"type":types.get(type_id,"Inconnu"),"label":label,"total_gb":total_gb}

def disk_watcher_loop():
    connus = lecteurs_presents()
    while True:
        try:
            actuels = lecteurs_presents()
            for lettre in actuels - connus:
                try:
                    infos = infos_lecteur(lettre)
                    log.info(f"Nouveau disque : {lettre}:")
                    state.nouveau_disque = infos
                    state.status = f"Disque branché : {lettre}: {infos['label']}"
                except Exception as e:
                    log.error(f"Erreur infos {lettre}: {e}")
            connus = actuels
        except Exception as e:
            log.error(f"DiskWatcher: {e}")
        time.sleep(DISK_CHECK_INTERVAL)

# ─── Watchdog Downloads ───────────────────────────────────────────────────────

class WatchHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory: return
        chemin = Path(event.src_path)
        if chemin.suffix.lower() not in TOUTES_EXT: return
        time.sleep(3)
        if not chemin.exists() or chemin.stat().st_size == 0: return
        taille_mb = round(chemin.stat().st_size / (1024*1024), 2)
        type_f    = type_fichier(chemin)
        serie, saison, episode = None, None, None
        if type_f == "Vidéo": serie, saison, episode = detecter_serie(chemin.name)
        inserer_fichier(chemin.name, type_f, str(DOWNLOADS_FOLDER.anchor).rstrip("\\/"),
                        chemin, taille_mb, serie, saison, episode, chemin.stem)
        state.status = f"Nouveau fichier : {chemin.name}"
        state.nouveau_fichier = True
        log.info(f"Downloads : {chemin.name}")

# ─── FastAPI ──────────────────────────────────────────────────────────────────

app = FastAPI(title="Media Indexer API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

class ScanRequest(BaseModel):
    path: str

class ExportRequest(BaseModel):
    path: str
    type_filtre:   Optional[str] = "Tous"
    disque_filtre: Optional[str] = "Tous"

@app.get("/status")
def get_status():
    nf = state.nouveau_fichier
    nd = state.nouveau_disque
    state.nouveau_fichier = False
    state.nouveau_disque  = None
    return {"message": state.status, "nouveau_fichier": nf, "nouveau_disque": nd}

@app.get("/fichiers")
def get_fichiers(recherche: str="", type_filtre: str="Tous", disque_filtre: str="Tous"):
    con = get_con()
    conds, params = [], []
    if recherche:
        q = f"%{recherche}%"
        conds.append("(nom LIKE ? OR serie LIKE ? OR titre LIKE ? OR tags LIKE ? OR resume LIKE ?)")
        params.extend([q,q,q,q,q])
    if type_filtre != "Tous": conds.append("type=?"); params.append(type_filtre)
    if disque_filtre != "Tous": conds.append("disque=?"); params.append(disque_filtre)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows = con.execute(
        f"SELECT id,date_scan,nom,type,disque,taille_mb,serie,saison,episode,titre,tags "
        f"FROM fichiers {where} ORDER BY serie,saison,episode,nom LIMIT 2000", params
    ).fetchall()
    con.close()
    keys = ["id","date_scan","nom","type","disque","taille_mb","serie","saison","episode","titre","tags"]
    return {"fichiers": [dict(zip(keys,r)) for r in rows]}

@app.get("/series")
def get_series(recherche: str=""):
    con = get_con()
    rows = con.execute("""
        SELECT serie,COUNT(*) as nb,GROUP_CONCAT(DISTINCT disque),MIN(saison),MAX(saison)
        FROM fichiers WHERE serie IS NOT NULL AND serie!=''
        GROUP BY serie ORDER BY serie
    """).fetchall()
    con.close()
    result = []
    for r in rows:
        if recherche and recherche.lower() not in (r[0] or "").lower(): continue
        result.append({"serie":r[0],"nb":r[1],"disques":r[2],"saison_min":r[3],"saison_max":r[4]})
    return {"series": result}

@app.get("/stats")
def get_stats():
    con      = get_con()
    total    = con.execute("SELECT COUNT(*) FROM fichiers").fetchone()[0]
    par_type = con.execute("SELECT type,COUNT(*),ROUND(SUM(taille_mb),0) FROM fichiers GROUP BY type ORDER BY COUNT(*) DESC").fetchall()
    disques  = con.execute("SELECT disque,COUNT(*) FROM fichiers GROUP BY disque ORDER BY COUNT(*) DESC").fetchall()
    con.close()
    return {
        "total":    total,
        "par_type": [{"type":r[0],"count":r[1],"taille_mb":r[2] or 0} for r in par_type],
        "disques":  [{"disque":r[0],"count":r[1]} for r in disques]
    }

@app.get("/detail/{fid}")
def get_detail(fid: int):
    con = get_con()
    row = con.execute(
        "SELECT id,nom,chemin,type,disque,taille_mb,serie,saison,episode,titre,resume,tags,date_scan FROM fichiers WHERE id=?",
        (fid,)
    ).fetchone()
    con.close()
    if not row: return JSONResponse(status_code=404, content={"error":"not found"})
    keys = ["id","nom","chemin","type","disque","taille_mb","serie","saison","episode","titre","resume","tags","date_scan"]
    return dict(zip(keys, row))

@app.get("/disques")
def get_disques():
    con  = get_con()
    rows = con.execute("SELECT DISTINCT disque FROM fichiers ORDER BY disque").fetchall()
    con.close()
    return {"disques": [r[0] for r in rows]}

@app.post("/scan")
def post_scan(req: ScanRequest):
    if state.scan_running:
        return {"message": "Scan déjà en cours."}
    threading.Thread(target=scanner_dossier, args=(req.path,), daemon=True).start()
    return {"message": f"Scan lancé : {req.path}"}

@app.post("/analyser/{fid}")
def post_analyser(fid: int):
    con = get_con()
    row = con.execute("SELECT chemin FROM fichiers WHERE id=?", (fid,)).fetchone()
    con.close()
    if not row: return JSONResponse(status_code=404, content={"error":"not found"})
    chemin = row[0]
    def run():
        state.status = f"Analyse Ollama…"
        titre, resume, tags = analyser_pdf_ollama(chemin)
        mettre_a_jour_analyse(chemin, titre, resume, tags)
        state.status = f"Analysé : {Path(chemin).name}"
        state.nouveau_fichier = True
    threading.Thread(target=run, daemon=True).start()
    return {"message": "Analyse lancée."}

@app.post("/export")
def post_export(req: ExportRequest):
    try:
        exporter_excel(req.path, req.type_filtre, req.disque_filtre)
        return {"message": f"Exporté : {req.path}"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    # Watcher Downloads
    handler  = WatchHandler()
    observer = Observer()
    observer.schedule(handler, str(DOWNLOADS_FOLDER), recursive=False)
    observer.daemon = True
    observer.start()
    # Disk watcher
    threading.Thread(target=disk_watcher_loop, daemon=True).start()
    log.info(f"Backend démarré sur http://{HOST}:{PORT}")
    uvicorn.run(app, host=HOST, port=PORT, log_level="warning")
